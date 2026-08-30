"""Import ward and polling-unit records from the supplied INEC-style workbook."""

import argparse
from pathlib import Path

from openpyxl import load_workbook

from app.db.session import SessionLocal
from app.models.location import PollingUnit, Ward


def clean(value):
    return str(value).strip() if value is not None else ""


def workbook_records(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    carried_ward = ""
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if clean(row[0]):
            carried_ward = clean(row[0])
        # The supplied sheets have no header and use columns A:G:
        # ward, sequence, state, LGA, ward, polling-unit name, polling-unit code.
        if len(row) < 7 or row[1] is None:
            continue
        state, lga, row_ward, name, code = map(clean, (row[2], row[3], row[4], row[5], row[6]))
        ward_name = row_ward or carried_ward
        if not all((lga, ward_name, name, code)):
            raise ValueError(f"Row {row_number} is missing location data; nothing was imported.")
        try:
            sequence_no = int(row[1])
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Row {row_number} has an invalid polling-unit sequence.") from exc
        yield state, lga, ward_name, name, code, sequence_no


def import_file(path: Path, target_members: int, dry_run: bool) -> tuple[int, int]:
    records = list(workbook_records(path))
    if not records:
        raise ValueError(f"{path} contains no polling-unit records.")

    db = SessionLocal()
    created = updated = 0
    try:
        for state, lga, ward_name, name, code, sequence_no in records:
            ward = db.query(Ward).filter(Ward.lga == lga, Ward.name == ward_name).first()
            if not ward:
                ward = Ward(state=state or None, lga=lga, name=ward_name)
                db.add(ward)
                db.flush()

            unit = db.query(PollingUnit).filter(PollingUnit.code == code).first()
            if not unit:
                db.add(PollingUnit(
                    ward_id=ward.id,
                    code=code,
                    name=name,
                    sequence_no=sequence_no,
                    target_members=target_members,
                ))
                created += 1
            else:
                # Update only data sourced from the workbook.  Preserve targets
                # that an administrator may already have customised.
                unit.ward_id = ward.id
                unit.name = name
                unit.sequence_no = sequence_no
                updated += 1

        if dry_run:
            db.rollback()
        else:
            db.commit()
        return created, updated
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path, nargs="+", help="One or more .xlsx files")
    parser.add_argument("--target", type=int, default=200, help="Initial target for newly imported units")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    if args.target < 0:
        parser.error("--target cannot be negative")
    for workbook in args.workbook:
        created, updated = import_file(workbook, args.target, args.dry_run)
        print(f"{workbook.name}: {created} created, {updated} updated" + (" (dry run)" if args.dry_run else ""))


if __name__ == "__main__":
    main()
