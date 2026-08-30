import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.config import APP_SHORT_NAME, APP_NAME


def generate_volunteers_excel(volunteers):
    os.makedirs("uploads", exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "AUM Volunteers"

    headers = [
        "ID",
        "Registration No",
        "Full Name",
        "Phone Number",
        "Gender",
        "Age",
        "LGA",
        "Ward",
        "Unit",
        "Highest Qualification",
        "Additional Qualification",
        "Specialization",
        "Employment Status",
        "Physically Challenged",
        "AUM Member",
        "Previous Organization",
        "Position",
        "Expectation",
        "Joined",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="0B2347",
        end_color="0B2347",
        fill_type="solid",
    )
    header_alignment = Alignment(horizontal="center")

    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for volunteer in volunteers:
        sheet.append([
            volunteer.id,
            volunteer.registration_no,
            volunteer.name,
            volunteer.phone,
            volunteer.gender,
            volunteer.age,
            volunteer.lga,
            volunteer.ward,
            volunteer.unit,
            volunteer.highest_qualification,
            volunteer.additional_qualification,
            volunteer.specialization,
            volunteer.employment_status,
            "Yes" if volunteer.physically_challenged else "No",
            "Yes" if volunteer.aum_member else "No",
            volunteer.previous_organization,
            volunteer.position,
            volunteer.expectation,
            volunteer.created_at.strftime("%d %B %Y") if volunteer.created_at else "",
        ])

    for column in sheet.columns:
        max_length = max(
            (len(str(cell.value)) for cell in column if cell.value is not None),
            default=0,
        )
        sheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 35)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = f"uploads/{APP_SHORT_NAME.lower()}_volunteers_{timestamp}.xlsx"
    workbook.properties.title = f"{APP_NAME} Volunteer Records"
    workbook.save(file_path)

    return file_path
